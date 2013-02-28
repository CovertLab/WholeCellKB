'''
Whole-cell helper functions

Author: Jonathan Karr, jkarr@stanford.edu
Affiliation: Covert Lab, Department of Bioengineering, Stanford University
Last updated: 2012-07-17
'''

from __future__ import unicode_literals
from BeautifulSoup import BeautifulStoneSoup
from copy import deepcopy
from dateutil.tz import tzlocal
from django.contrib.auth.models import User
from django.core import serializers
from django.core.exceptions import ValidationError, ObjectDoesNotExist
from django.core.urlresolvers import reverse
from django.db import IntegrityError
from django.db.models import get_app, get_models, get_model
from django.db.models.fields import AutoField, BigIntegerField, IntegerField, PositiveIntegerField, PositiveSmallIntegerField, SmallIntegerField, BooleanField, NullBooleanField, DecimalField, FloatField, CharField, CommaSeparatedIntegerField, EmailField, FilePathField, GenericIPAddressField, IPAddressField, SlugField, URLField, TextField, DateField, DateTimeField, TimeField, NOT_PROVIDED
from django.db.models.fields.related import OneToOneField, RelatedObject, ManyToManyField, ForeignKey
from django.db.models.query import EmptyQuerySet
from django.http import Http404, HttpResponse, HttpResponseBadRequest
from django.shortcuts import render_to_response, get_object_or_404
from django.template import Context, RequestContext, loader
from django.template.loader import get_template
from django.utils import simplejson
from django.utils.html import strip_tags
from django.utils.text import capfirst
from odict import odict
from openpyxl import Workbook, load_workbook
from openpyxl.cell import Cell, get_column_letter
from openpyxl.shared.date_time import SharedDate, CALENDAR_WINDOWS_1900
from openpyxl.style import NumberFormat, Border, Color, HashableObject, Protection, Alignment
from openpyxl.writer.styles import StyleWriter
from public.templatetags.templatetags import ceil
from public.models import Entry, Species, SpeciesComponent, Reference, Chromosome, format_list_html, format_evidence, Evidence, EvidencedEntryData
from public.models import EntryData, EntryBooleanData, EntryCharData, EntryFloatData, EntryPositiveFloatData, EntryTextData
from StringIO import StringIO
from subprocess import Popen, PIPE
from xml.dom.minidom import Document
import datetime
import htmlentitydefs
import inspect
import math
import os
import re
import settings
import sys
import tempfile
import time
import xhtml2pdf.pisa as pisa

class ELEMENT_MWS:
	H = 1.0079
	He = 4.0026
	Li = 6.941
	Be = 9.0122
	B = 10.811
	C = 12.0107
	N = 14.0067
	O = 15.9994
	F = 18.9984
	Ne = 20.1797
	Na = 22.9897
	Mg = 24.305
	Al = 26.9815
	Si = 28.0855
	P = 30.9738
	S = 32.065
	Cl = 35.453
	Ar = 39.948
	K = 39.0983
	Ca = 40.078
	Sc = 44.9559
	Ti = 47.867
	V = 50.9415
	Cr = 51.9961
	Mn = 54.938
	Fe = 55.845
	Co = 58.9332
	Ni = 58.6934
	Cu = 63.546
	Zn = 65.39
	Ga = 69.723
	Ge = 72.64
	As = 74.9216
	Se = 78.96
	Br = 79.904
	Kr = 83.8
	Rb = 85.4678
	Sr = 87.62
	Y = 88.9059
	Zr = 91.224
	Nb = 92.9064
	Mo = 95.94
	Tc = 98
	Ru = 101.07
	Rh = 102.9055
	Pd = 106.42
	Ag = 107.8682
	Cd = 112.411
	In = 114.818
	Sn = 118.71
	Sb = 121.76
	Te = 127.6
	I = 126.9045
	Xe = 131.293
	Cs = 132.9055
	Ba = 137.327
	La = 138.9055
	Ce = 140.116
	Pr = 140.9077
	Nd = 144.24
	Pm = 145
	Sm = 150.36
	Eu = 151.964
	Gd = 157.25
	Tb = 158.9253
	Dy = 162.5
	Ho = 164.9303
	Er = 167.259
	Tm = 168.9342
	Yb = 173.04
	Lu = 174.967
	Hf = 178.49
	Ta = 180.9479
	W = 183.84
	Re = 186.207
	Os = 190.23
	Ir = 192.217
	Pt = 195.078
	Au = 196.9665
	Hg = 200.59
	Tl = 204.3833
	Pb = 207.2
	Bi = 208.9804
	Po = 209
	At = 210
	Rn = 222
	Fr = 223
	Ra = 226
	Ac = 227
	Th = 232.0381
	Pa = 231.0359
	U = 238.0289
	Np = 237
	Pu = 244
	Am = 243
	Cm = 247
	Bk = 247
	Cf = 251
	Es = 252
	Fm = 257
	Md = 258
	No = 259
	Lr = 262
	Rf = 261
	Db = 262
	Sg = 266
	Bh = 264
	Hs = 277
	Mt = 268
	
#Source: http://www.owczarzy.net/extinct.htm
class ExtinctionCoefficient:
	single_dna = {
		'A': 15400,
		'C':  7200,
		'G': 11500,
		'T':  8700,
		}
	single_rna = {
		'A': 15400,
		'C':  7200,
		'G': 11500,
		'U':  9900,
		}
	pairwise_dna = {
		'A': {'A': 27400, 'C': 21200, 'G': 25000, 'T': 28000},
		'C': {'A': 21200, 'C': 14600, 'G': 18000, 'T': 15200},
		'G': {'A': 25200, 'C': 17600, 'G': 21600, 'T': 20000},
		'T': {'A': 23400, 'C': 16200, 'G': 19000, 'T': 16800},
		}
	pairwise_rna = {
		'A': {'A': 27400, 'C': 21000, 'G': 25000, 'U': 24000},
		'C': {'A': 21000, 'C': 14200, 'G': 17800, 'U': 16200},
		'G': {'A': 25200, 'C': 17400, 'G': 21600, 'U': 21200},
		'U': {'A': 24600, 'C': 17200, 'G': 20000, 'U': 19600},
		}

'''
dipeptide instability weight value (DIWV)
1. http://ca.expasy.org/tools/protparam-doc.html
2. Kunchur Guruprasad, B.V.Bhasker Reddy and Madhusudan W.Pandit
   (1990). Correlation between stability of a protein and its
   dipeptide composition: a novel approach for predicting in vivo
   stability of a protein from its primary sequence. Protein
   Engineering 4(2):155-161.
'''
class DipeptideInstabilityWeight:
	value = {
	"A": {"A":   1.00, "R":   1.00, "N":   1.00, "D":  -7.49, "C": 44.94, "Q":  1.00, "E":  1.00, "G":   1.00, "H": -7.49, "I":  1.00, "L":  1.00, "K":   1.00, "M":  1.00, "F":   1.00, "P": 20.26, "S":  1.00, "T":   1.00, "W":   1.00, "Y":  1.00, "V":  1.00, },
	"R": {"A":   1.00, "R":  58.28, "N":  13.34, "D":   1.00, "C":  1.00, "Q": 20.26, "E":  1.00, "G":  -7.49, "H": 20.26, "I":  1.00, "L":  1.00, "K":   1.00, "M":  1.00, "F":   1.00, "P": 20.26, "S": 44.94, "T":   1.00, "W":  58.28, "Y": -6.54, "V":  1.00, },
	"N": {"A":   1.00, "R":   1.00, "N":   1.00, "D":   1.00, "C": -1.88, "Q": -6.54, "E":  1.00, "G": -14.03, "H":  1.00, "I": 44.94, "L":  1.00, "K":  24.68, "M":  1.00, "F": -14.03, "P": -1.88, "S":  1.00, "T":  -7.49, "W":  -9.37, "Y":  1.00, "V":  1.00, },
	"D": {"A":   1.00, "R":  -6.54, "N":   1.00, "D":   1.00, "C":  1.00, "Q":  1.00, "E":  1.00, "G":   1.00, "H":  1.00, "I":  1.00, "L":  1.00, "K":  -7.49, "M":  1.00, "F":  -6.54, "P":  1.00, "S": 20.26, "T": -14.03, "W":   1.00, "Y":  1.00, "V":  1.00, },
	"C": {"A":   1.00, "R":   1.00, "N":   1.00, "D":  20.26, "C":  1.00, "Q": -6.54, "E":  1.00, "G":   1.00, "H": 33.60, "I":  1.00, "L": 20.26, "K":   1.00, "M": 33.60, "F":   1.00, "P": 20.26, "S":  1.00, "T":  33.60, "W":  24.68, "Y":  1.00, "V": -6.54, },
	"Q": {"A":   1.00, "R":   1.00, "N":   1.00, "D":  20.26, "C": -6.54, "Q": 20.26, "E": 20.26, "G":   1.00, "H":  1.00, "I":  1.00, "L":  1.00, "K":   1.00, "M":  1.00, "F":  -6.54, "P": 20.26, "S": 44.94, "T":   1.00, "W":   1.00, "Y": -6.54, "V": -6.54, },
	"E": {"A":   1.00, "R":   1.00, "N":   1.00, "D":  20.26, "C": 44.94, "Q": 20.26, "E": 33.60, "G":   1.00, "H": -6.54, "I": 20.26, "L":  1.00, "K":   1.00, "M":  1.00, "F":   1.00, "P": 20.26, "S": 20.26, "T":   1.00, "W": -14.03, "Y":  1.00, "V":  1.00, },
	"G": {"A":  -7.49, "R":   1.00, "N":  -7.49, "D":   1.00, "C":  1.00, "Q":  1.00, "E": -6.54, "G":  13.34, "H":  1.00, "I": -7.49, "L":  1.00, "K":  -7.49, "M":  1.00, "F":   1.00, "P":  1.00, "S":  1.00, "T":  -7.49, "W":  13.34, "Y": -7.49, "V":  1.00, },
	"H": {"A":   1.00, "R":   1.00, "N":  24.68, "D":   1.00, "C":  1.00, "Q":  1.00, "E":  1.00, "G":  -9.37, "H":  1.00, "I": 44.94, "L":  1.00, "K":  24.68, "M":  1.00, "F":  -9.37, "P": -1.88, "S":  1.00, "T":  -6.54, "W":  -1.88, "Y": 44.94, "V":  1.00, },
	"I": {"A":   1.00, "R":   1.00, "N":   1.00, "D":   1.00, "C":  1.00, "Q":  1.00, "E": 44.94, "G":   1.00, "H": 13.34, "I":  1.00, "L": 20.26, "K":  -7.49, "M":  1.00, "F":   1.00, "P": -1.88, "S":  1.00, "T":   1.00, "W":   1.00, "Y":  1.00, "V": -7.49, },
	"L": {"A":   1.00, "R":  20.26, "N":   1.00, "D":   1.00, "C":  1.00, "Q": 33.60, "E":  1.00, "G":   1.00, "H":  1.00, "I":  1.00, "L":  1.00, "K":  -7.49, "M":  1.00, "F":   1.00, "P": 20.26, "S":  1.00, "T":   1.00, "W":  24.68, "Y":  1.00, "V":  1.00, },
	"K": {"A":   1.00, "R":  33.60, "N":   1.00, "D":   1.00, "C":  1.00, "Q": 24.68, "E":  1.00, "G":  -7.49, "H":  1.00, "I": -7.49, "L": -7.49, "K":   1.00, "M": 33.60, "F":   1.00, "P": -6.54, "S":  1.00, "T":   1.00, "W":   1.00, "Y":  1.00, "V": -7.49, },
	"M": {"A":  13.34, "R":  -6.54, "N":   1.00, "D":   1.00, "C":  1.00, "Q": -6.54, "E":  1.00, "G":   1.00, "H": 58.28, "I":  1.00, "L":  1.00, "K":   1.00, "M": -1.88, "F":   1.00, "P": 44.94, "S": 44.94, "T":  -1.88, "W":   1.00, "Y": 24.68, "V":  1.00, },
	"F": {"A":   1.00, "R":   1.00, "N":   1.00, "D":  13.34, "C":  1.00, "Q":  1.00, "E":  1.00, "G":   1.00, "H":  1.00, "I":  1.00, "L":  1.00, "K": -14.03, "M":  1.00, "F":   1.00, "P": 20.26, "S":  1.00, "T":   1.00, "W":   1.00, "Y": 33.60, "V":  1.00, },
	"P": {"A":  20.26, "R":  -6.54, "N":   1.00, "D":  -6.54, "C": -6.54, "Q": 20.26, "E": 18.38, "G":   1.00, "H":  1.00, "I":  1.00, "L":  1.00, "K":   1.00, "M": -6.54, "F":  20.26, "P": 20.26, "S": 20.26, "T":   1.00, "W":  -1.88, "Y":  1.00, "V": 20.26, },
	"S": {"A":   1.00, "R":  20.26, "N":   1.00, "D":   1.00, "C": 33.60, "Q": 20.26, "E": 20.26, "G":   1.00, "H":  1.00, "I":  1.00, "L":  1.00, "K":   1.00, "M":  1.00, "F":   1.00, "P": 44.94, "S": 20.26, "T":   1.00, "W":   1.00, "Y":  1.00, "V":  1.00, },
	"T": {"A":   1.00, "R":   1.00, "N": -14.03, "D":   1.00, "C":  1.00, "Q": -6.54, "E": 20.26, "G":  -7.49, "H":  1.00, "I":  1.00, "L":  1.00, "K":   1.00, "M":  1.00, "F":  13.34, "P":  1.00, "S":  1.00, "T":   1.00, "W": -14.03, "Y":  1.00, "V":  1.00, },
	"W": {"A": -14.03, "R":   1.00, "N":  13.34, "D":   1.00, "C":  1.00, "Q":  1.00, "E":  1.00, "G":  -9.37, "H": 24.68, "I":  1.00, "L": 13.34, "K":   1.00, "M": 24.68, "F":   1.00, "P":  1.00, "S":  1.00, "T": -14.03, "W":   1.00, "Y":  1.00, "V": -7.49, },
	"Y": {"A":  24.68, "R": -15.91, "N":   1.00, "D":  24.68, "C":  1.00, "Q":  1.00, "E": -6.54, "G":  -7.49, "H": 13.34, "I":  1.00, "L":  1.00, "K":   1.00, "M": 44.94, "F":   1.00, "P": 13.34, "S":  1.00, "T":  -7.49, "W":  -9.37, "Y": 13.34, "V":  1.00, },
	"V": {"A":   1.00, "R":   1.00, "N":   1.00, "D": -14.03, "C":  1.00, "Q":  1.00, "E":  1.00, "G":  -7.49, "H":  1.00, "I":  1.00, "L":  1.00, "K":  -1.88, "M":  1.00, "F":   1.00, "P": 20.26, "S":  1.00, "T":  -7.49, "W":   1.00, "Y": -6.54, "V":  1.00, },
	} 
	
class Colors(HashableObject):
	BLACK = 'FF000000'
	BLUE = 'FF306EFF'
	GREY1 = 'FFCCCCCC'
	GREY2 = 'FFDDDDDD'
	GREY3 = 'FF999999'
	GREY4 = 'FFAAAAAA'
	WHITE = 'FFFFFFFF'

	__fields__ = ('index',)
	__slots__ = __fields__
	__leaf__ = True

	def __init__(self, index):
		super(Color, self).__init__()
		self.index = index

def getModels(superclass=Entry):
	tmp = get_models(get_app('public'))
	tmp2 = {}
	for model in tmp:
		if issubclass(model, superclass) and model._meta.concrete_entry_model:
			tmp2[model.__name__] = model
	return tmp2

def get_models_in_saving_order(superclass=Entry):
	all_models = get_models(get_app('public'))
	entry_models = []
	for model in all_models:
		if issubclass(model, superclass) and model._meta.concrete_entry_model:
			entry_models.append(model)

	return sorted(entry_models, cmp=model_save_order_comparator)

def model_save_order_comparator(x, y):
	if is_model_referenced(x, y):
		return 1

	if is_model_referenced(y, x):
		return -1

	return 0

def is_model_referenced(model, referenced_model, checked_models=[]):
	for field in model._meta.fields + model._meta.many_to_many:
		if isinstance(field, (ForeignKey, ManyToManyField)) and not (isinstance(field, OneToOneField) and field.rel.parent_link) and issubclass(field.rel.to, Entry):
			if issubclass(referenced_model, field.rel.to):
				return True

	for field in model._meta.fields + model._meta.many_to_many:
		if isinstance(field, (ForeignKey, ManyToManyField)) and not (isinstance(field, OneToOneField) and field.rel.parent_link) and issubclass(field.rel.to, Entry):
			if (field.rel.to not in checked_models) and is_model_referenced(field.rel.to, referenced_model, checked_models + [model]):
				return True
	return False

def getModelsMetadata(superclass=Entry):
	tmp = get_models(get_app('public'))
	tmp2 = {}
	for model in tmp:
		if issubclass(model, superclass) and model._meta.concrete_entry_model:
			tmp2[model.__name__] = model._meta
	return tmp2

def getObjectTypes(superclass=Entry):
	tmp = get_models(get_app('public'))
	tmp2 = []
	for model in tmp:
		if issubclass(model, superclass) and model._meta.concrete_entry_model:
			tmp2.append(model.__name__)
	tmp2.sort()
	return tmp2

def getModel(str):
	tmp = get_models(get_app('public'))
	tmp2 = []
	for model in tmp:
		if issubclass(model, Entry) and model._meta.concrete_entry_model and model.__name__ == str:
			return model
	return

def getModelByVerboseName(str):
	tmp = get_models(get_app('public'))
	tmp2 = []
	for model in tmp:
		if issubclass(model, Entry) and model._meta.concrete_entry_model and model.verbose_name == str:
			return model
	return

def getModelByVerboseNamePlural(str):
	tmp = get_models(get_app('public'))
	tmp2 = []
	for model in tmp:
		if issubclass(model, Entry) and model._meta.concrete_entry_model and model._meta.verbose_name_plural == str:
			return model
	return

def getModelAdmin(str):
	app_ = __import__('public')
	admins_ = getattr(app_, 'admin')
	if hasattr(admins_, str + 'Admin'):
		return getattr(admins_, str + 'Admin')
	return None

def getModelDataFields(model, auto_created=True, metadata=True):
	unordered_fields = []
	for field in model._meta.fields + model._meta.many_to_many:
		if not (field.name in ['species', 'model_type'] or (isinstance(field, OneToOneField) and field.rel.parent_link)) and (not field.auto_created or auto_created):
			unordered_fields.append(field)
			
	ordered_fields = []
	for field in model._meta.field_list:
		if not metadata and field in ['id', 'created_user', 'last_updated_user']:
			continue
		
		field = model._meta.get_field_by_name(field)[0]
		if not field.auto_created or auto_created:
			ordered_fields.append(field)
	
	if metadata and (len(unordered_fields) != len(ordered_fields) or set(unordered_fields) != set(ordered_fields)):
		print str(set(unordered_fields).difference(ordered_fields))
		print str(set(ordered_fields).difference(unordered_fields))
		raise Http404
	
	return ordered_fields

def getModelDataFieldNames(model):
	fields = []
	for field in model._meta.fields + model._meta.many_to_many:
		if not (field.name == 'model_type' or (isinstance(field, OneToOneField) and field.rel.parent_link)):
			fields.append(field.name)
	return fields

def is_entrydata_model(cls):
	return inspect.isclass(cls) and issubclass(cls, EntryData)

def getEntryDatas():
	return inspect.getmembers(sys.modules['public.models'], is_entrydata_model)

def getEntry(species_wid = None, wid = None):
	try:
		species_id = Species.objects.values('id').get(wid = species_wid)['id']
	except ObjectDoesNotExist:
		return None	
	
	try:
		if species_wid == wid:
			return Species.objects.get(wid=wid)
		else:
			tmp = SpeciesComponent.objects.get(species__id = species_id, wid=wid)
			return getModel(tmp.model_type).objects.select_related(depth=2).get(id=tmp.id)
	except ObjectDoesNotExist:
		return None
	
def html_to_ascii(s):
	return strip_tags(unicode(BeautifulStoneSoup(s, convertEntities=BeautifulStoneSoup.ALL_ENTITIES)).replace("&nbsp;", " "))

'''
Element order preserved
From http://www.peterbe.com/plog/uniqifiers-benchmark
'''
def uniqueUnsorted(seq):
	seen = set()
	return [x for x in seq if x not in seen and not seen.add(x)]

'''
Element order not preserved
From http://www.peterbe.com/plog/uniqifiers-benchmark
'''
def uniqueSorted(seq):
	return {}.fromkeys(seq).keys()

def render_queryset_to_response(request = [], queryset = EmptyQuerySet(), models = [], templateFile = '', data = {}, species_wid=None):
	format = request.GET.get('format', 'html')
	
	if species_wid is not None and species_wid != '':
		species = Species.objects.get(wid=species_wid)
	else:
		species = Species.objects.all()
		if len(species) > 0:
			species = species[0]
		else:
			species = None
	
	data['species'] = species
	data['queryset'] = queryset
	data['request'] = request
	data['queryargs'] = {}
	for key, val in request.GET.iterlists():
		data['queryargs'][key] = val

	if format == 'html':
		data['is_pdf'] = False
		data['pdfstyles'] = ''
		data['species_list'] = Species.objects.all()
		data['modelmetadatas'] = getModelsMetadata(SpeciesComponent)
		data['modelnames'] = getObjectTypes(SpeciesComponent)
		data['last_updated_date'] = datetime.datetime.fromtimestamp(os.path.getmtime(settings.TEMPLATE_DIRS[0] + '/' + templateFile))
		
		return render_to_response(templateFile, data, context_instance = RequestContext(request))
	elif format == 'bib':
		response = HttpResponse(
			write_bibtex(species, queryset),
			mimetype = "application/x-bibtex; charset=UTF-8",
			content_type = "application/x-bibtex; charset=UTF-8")
		response['Content-Disposition'] = "attachment; filename=data.bib"
	elif format == 'json':
		objects = []
		for obj in queryset.iterator():
			objDict = convert_modelobject_to_stdobject(obj, request.user.is_anonymous())
			objDict['model'] = obj.__class__.__name__
			objects.append(objDict)
		
		now = datetime.datetime.now(tzlocal())		
		json = odict()
		json['title'] = '%s WholeCellKB' % species.name
		json['comments'] = 'Generated by %s on %s at %s' % ('WholeCellKB', now.isoformat(), settings.ROOT_URL + reverse('public.views.exportData', kwargs={'species_wid': species.wid}))
		json['copyright'] = '%s %s' % (now.year, 'Covert Lab, Department of Bioengineering, Stanford University')
		json['data'] = objects
		response = HttpResponse(
			simplejson.dumps(json, indent=2, ensure_ascii=False, encoding='utf-8'),
			mimetype = "application/json; charset=UTF-8",
			content_type = "application/json; charset=UTF-8")
		response['Content-Disposition'] = "attachment; filename=data.json"
	elif format == 'pdf':
		data['is_pdf'] = True
		data['pdfstyles'] = ''
		data['species_list'] = Species.objects.all()
		data['modelmetadatas'] = getModelsMetadata(SpeciesComponent)
		data['modelnames'] = getObjectTypes(SpeciesComponent)

		for fileName in ['styles', 'styles.print', 'styles.pdf']:
			f = open(settings.STATICFILES_DIRS[0] + '/public/css/' + fileName + '.css', 'r')
			data['pdfstyles'] += f.read()
			f.close()

		response = HttpResponse(
			mimetype = 'application/pdf',
			content_type = 'application/pdf')
		response['Content-Disposition'] = "attachment; filename=data.pdf"

		template = get_template(templateFile)

		pdf = pisa.CreatePDF(
			src = template.render(Context(data)),
			dest = response)

		if not pdf.err:
			return response
		return Http404
	elif format == 'xlsx':
		#write work book
		wb = writeExcel(species, queryset, models, request.user.is_anonymous())

		#save to string
		result = StringIO()
		wb.save(filename = result)

		#generate HttpResponse
		response = HttpResponse(
			result.getvalue(),
			mimetype = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
			content_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
		response['Content-Disposition'] = "attachment; filename=data.xlsx"
	elif format == 'xml':
		doc = Document()
		
		now = datetime.datetime.now(tzlocal())
		comment = doc.createComment('\n%s WholeCellKB\nGenerated by %s on %s at %s\n%s %s %s\n' % (
			species.name, 
			'WholeCellKB', now.isoformat(), 
			settings.ROOT_URL + reverse('public.views.exportData', kwargs={'species_wid': species.wid}),
			html_to_ascii('&copy;'), now.year, 'Covert Lab, Department of Bioengineering, Stanford University',
			))
		doc.appendChild(comment)
		
		objects = doc.createElement('objects')
		doc.appendChild(objects)

		for obj in queryset:
			objects.appendChild(convert_modelobject_to_xml(obj, doc, request.user.is_anonymous()))

		response = HttpResponse(
			doc.toprettyxml(indent=" "*2, encoding = 'utf-8'),
			mimetype = "application/xml; charset=UTF-8",
			content_type = "application/xml; charset=UTF-8")
		response['Content-Disposition'] = "attachment; filename=data.xml"
	else:
		t = loader.get_template('public/error.html')
		data['type'] = 500
		data['message'] = '"%s" is not a supported export format.' % format
		c = Context(data)
		response =  HttpResponseBadRequest(
			t.render(c),
			mimetype = 'text/html; charset=UTF-8',
			content_type = 'text/html; charset=UTF-8')

	return response

def writeExcel(species, queryset, modelArr, is_user_anonymous):
	#sort entry models by name
	modelDict = {}
	model_types = []
	objectDict = {}
	for model in modelArr:
		model_types.append(model.__name__)
		modelDict[model.__name__] = model
		objectDict[model.__name__] = []
	model_types.sort()

	#organize entries by model
	for obj in queryset:
		objectDict[obj.__class__.__name__].append(obj)
		
	#create workbook
	wb = Workbook()
	wb.remove_sheet(wb.get_active_sheet())

	#meta data
	now = datetime.datetime.now(tzlocal())
	wb.properties.creator = 'Generated by WholeCellKB on %s at %s' % (
		now.isoformat(), 
		settings.ROOT_URL + reverse('public.views.exportData', kwargs={'species_wid': species.wid}),
		)
	wb.properties.last_modified_by = wb.properties.creator
	wb.properties.created = now
	wb.properties.modified = now
	wb.properties.title = '%s WholeCellKB' % species.name
	wb.properties.subject = wb.properties.title
	wb.properties.description = wb.properties.title
	wb.properties.keywords = 'biology, systems, database, knowledge base'
	wb.properties.category = ''
	wb.properties.company = 'Covert Lab, Department of Bioengineering, Stanford University'
	wb.properties.excel_base_date = CALENDAR_WINDOWS_1900
	
	#print table of contents
	ws = wb.create_sheet(title = 'Contents')
	ws.default_column_dimension.auto_size = True
	ws.default_row_dimension.height = 15

	headers = (
		(wb.properties.title, 18, True),
		('Generated by WholeCellKB on %s at %s' % (
			now.strftime("%Y-%m-%d"), 
			settings.ROOT_URL + reverse('public.views.exportData', kwargs={'species_wid': species.wid}),
			), 10, False),
		('%s %s %s' % (html_to_ascii('&copy;'), now.year, wb.properties.company), 10, False),
	)	
	for i in range(len(headers)):		
		ws.merge_cells(start_row=i, start_column=0, end_row=i, end_column=1)
		cell = ws.cell(row=i, column=0)
		cell.set_value_explicit(value = headers[i][0], data_type = Cell.TYPE_STRING)	
		cell.style.number_format.format_code = NumberFormat.FORMAT_TEXT
		cell.style.font.name = 'Arial'
		cell.style.font.size = headers[i][1]
		cell.style.font.bold = headers[i][2]
		cell.style.alignment.wrap_text = True
		cell.style.alignment.horizontal = Alignment.HORIZONTAL_CENTER
		cell.style.fill.fill_type = 'solid'
		cell.style.fill.start_color.index = Colors.GREY1
		
		ws.row_dimensions[i + 1].height = headers[i][1] + 7

	cell = ws.cell(row=len(headers) + 1, column=0)
	cell.set_value_explicit(value = 'Table', data_type = Cell.TYPE_STRING)	
	cell.style.number_format.format_code = NumberFormat.FORMAT_TEXT
	cell.style.font.name = 'Arial'
	cell.style.font.size = 10
	cell.style.font.bold = True
	cell.style.alignment.wrap_text = True
	cell.style.alignment.horizontal = Alignment.HORIZONTAL_RIGHT
	cell.style.borders.bottom.color.index = Colors.BLACK
	cell.style.borders.bottom.border_style = Border.BORDER_THIN
	
	cell = ws.cell(row=len(headers) + 1, column=1)
	cell.set_value_explicit(value = 'Title', data_type = Cell.TYPE_STRING)	
	cell.style.number_format.format_code = NumberFormat.FORMAT_TEXT
	cell.style.font.name = 'Arial'
	cell.style.font.size = 10
	cell.style.font.bold = True
	cell.style.alignment.wrap_text = True
	cell.style.alignment.horizontal = Alignment.HORIZONTAL_LEFT
	cell.style.borders.bottom.color.index = Colors.BLACK
	cell.style.borders.bottom.border_style = Border.BORDER_THIN
	
	ws.freeze_panes = str('A%s' % (len(headers) + 3))
	ws.column_dimensions['A'].width = 8
	ws.column_dimensions['B'].width = 75
		
	for i in range(len(model_types)):
		iRow = i + len(headers) + 2
		model = modelDict[model_types[i]]
		
		cell = ws.cell(row=iRow, column=0)
		cell.set_value_explicit(value = '%d.' % (i+1), data_type = Cell.TYPE_STRING)	
		cell.style.number_format.format_code = NumberFormat.FORMAT_TEXT
		cell.style.font.name = 'Arial'
		cell.style.font.size = 10
		cell.style.alignment.wrap_text = True
		cell.style.alignment.horizontal = Alignment.HORIZONTAL_RIGHT
		cell.hyperlink = '#\'%s\'!%s' % (model._meta.verbose_name_plural[:31], 'A1') 
		
		cell = ws.cell(row=iRow, column=1)
		cell.set_value_explicit(value = model._meta.verbose_name_plural, data_type = Cell.TYPE_STRING)	
		cell.style.number_format.format_code = NumberFormat.FORMAT_TEXT
		cell.style.font.name = 'Arial'
		cell.style.font.size = 10
		cell.style.alignment.wrap_text = True
		cell.style.alignment.horizontal = Alignment.HORIZONTAL_LEFT		
		cell.hyperlink = '#\'%s\'!%s' % (model._meta.verbose_name_plural[:31], 'A1')

	#print kbojects to worksheets, 1 for each model
	for model_type in model_types:
		model = modelDict[model_type]
		fields = getModelDataFields(model, metadata=not is_user_anonymous)
		
		iCol = -1
		iCol_id = None
		iCol_wid = None
		for header, subheaders in get_excel_headers(model, is_user_anonymous=is_user_anonymous):
			if header == 'ID':
				iCol_id = iCol + 1
			elif header == 'WID':
				iCol_wid = iCol + 1
			iCol += max(1, len(subheaders))

		#create worksheet
		ws = wb.create_sheet(title = model._meta.verbose_name_plural[:31])
		ws.freeze_panes = str(get_column_letter(iCol_wid + 1) + '3')

		#header
		iRow1 = 0
		iRow2 = 1
		iCol = -1
		for header, subheaders in get_excel_headers(model, is_user_anonymous=is_user_anonymous):
			if len(subheaders) <= 1:
				iCol += 1
				cell = ws.cell(row=iRow2, column=iCol)
				cell.set_value_explicit(value = header, data_type = Cell.TYPE_STRING)
			else:
				ws.merge_cells(start_row=iRow1, start_column=iCol+1, end_row=iRow1, end_column=iCol+len(subheaders))
				cell = ws.cell(row=iRow1, column=iCol+1)
				cell.set_value_explicit(value = header, data_type = Cell.TYPE_STRING)
				for subheader in subheaders:
					iCol += 1
					cell = ws.cell(row=iRow2, column=iCol)
					cell.set_value_explicit(value = subheader, data_type = Cell.TYPE_STRING)

		#data
		iRow = 1
		for obj in objectDict[model_type]:
			iRow += 1
			iCol = -1
			for field in fields:
				value, data_type, format_code = format_value_excel(obj, field)				
				try:
					delattr(obj, "_" + field.name + "_cache")
				except Exception, e:
					pass

				if not isinstance(value, list):
					value = [value,]
					data_type = [data_type, ]
					format_code = [format_code, ]

				for iter in range(len(value)):
					iCol += 1
					cell = ws.cell(row=iRow, column=iCol)
					if value[iter] is None:
						cell.set_value_explicit(b'', data_type = Cell.TYPE_NULL)
					else:					
						cell.set_value_explicit(value = value[iter], data_type = data_type[iter])
					cell.style.number_format.format_code = format_code[iter]
					cell.style.font.name = 'Arial'
					cell.style.font.size = 10
					cell.style.alignment.wrap_text = True
					cell.style.fill.fill_type = 'solid'
					if iRow % 2 == 0:
						cell.style.fill.start_color.index = Colors.GREY1
					else:
						cell.style.fill.start_color.index = Colors.GREY2

		#style
		nCols = ws.get_highest_column()
		nRows = ws.get_highest_row()
		for iRow in range(nRows):
			cell = ws.cell(row=iRow, column=0)
			cell.style.font.bold = True
			if iRow > 1:
				if iRow % 2 == 0:
					cell.style.fill.start_color.index = Colors.GREY3
				else:
					cell.style.fill.start_color.index = Colors.GREY4
			if iRow > 0:
				cell.style.borders.left.color.index = Colors.BLUE
				cell.style.borders.left.border_style = Border.BORDER_THIN
				cell.style.borders.right.color.index = Colors.BLUE
				cell.style.borders.right.border_style = Border.BORDER_THIN

				cell = ws.cell(row=iRow, column=nCols-1)
				cell.style.borders.right.color.index = Colors.BLUE
				cell.style.borders.right.border_style = Border.BORDER_THIN

			ws.row_dimensions[iRow + 1].height = 15
		for iCol in range(nCols):
			cell = ws.cell(row=0, column=iCol)
			if cell.value is not None:
				cell.style.number_format.format_code = NumberFormat.FORMAT_TEXT
				cell.style.font.name = 'Arial'
				cell.style.font.size = 10
				cell.style.font.color.index = Colors.WHITE
				cell.style.font.bold = True
				cell.style.alignment.horizontal = Alignment.HORIZONTAL_CENTER
				cell.style.alignment.wrap_text = True
				cell.style.fill.fill_type = 'solid'
				cell.style.fill.start_color.index = Colors.BLUE
				cell.style.borders.top.color.index = Colors.BLUE
				cell.style.borders.top.border_style = Border.BORDER_THIN
				if (iCol == 0) or (ws.cell(row=0, column=iCol-1).value is None):
					cell.style.borders.left.color.index = Colors.BLUE
					cell.style.borders.left.border_style = Border.BORDER_THIN
				if (iCol == nCols - 1) or (ws.cell(row=0, column=iCol+1).value is None):
					cell.style.borders.right.color.index = Colors.BLUE
					cell.style.borders.right.border_style = Border.BORDER_THIN
			else:
				cell.style.borders.bottom.color.index = Colors.BLUE
				cell.style.borders.bottom.border_style = Border.BORDER_THIN

			cell = ws.cell(row=1, column=iCol)
			cell.style.number_format.format_code = NumberFormat.FORMAT_TEXT
			cell.style.font.name = 'Arial'
			cell.style.font.size = 10
			cell.style.font.color.index = Colors.WHITE
			cell.style.font.bold = True
			cell.style.alignment.wrap_text = True
			cell.style.fill.fill_type = 'solid'
			cell.style.fill.start_color.index = Colors.BLUE
			cell.style.borders.bottom.color.index = Colors.BLUE
			cell.style.borders.bottom.border_style = Border.BORDER_THIN

			cell = ws.cell(row=nRows-1, column=iCol)
			cell.style.borders.bottom.color.index = Colors.BLUE
			cell.style.borders.bottom.border_style = Border.BORDER_THIN

	return wb

def get_excel_headers(model, is_user_anonymous=False):
	fields = getModelDataFields(model, metadata=not is_user_anonymous)
	headers = []
	for field in fields:
		subheaders = []
		if (field.rel is not None) and isinstance(field, ForeignKey) and (not issubclass(field.rel.to, (Entry, User, ))):
			for subfield in field.rel.to._meta.fields + field.rel.to._meta.many_to_many:
				if not subfield.auto_created:
					subheaders.append(html_to_ascii(subfield.verbose_name))
		headers.append((html_to_ascii(field.verbose_name), subheaders, ))
	return headers
	
def format_value_excel(obj, field, depth = 0):
	value = None
	if hasattr(obj, field.name):
		value = getattr(obj, field.name)
	if field.__class__ in [AutoField, BigIntegerField, IntegerField, PositiveIntegerField, PositiveSmallIntegerField, SmallIntegerField]:
		data_type = Cell.TYPE_NUMERIC
		format_code = NumberFormat.FORMAT_NUMBER
	elif field.__class__ in [BooleanField, NullBooleanField]:
		data_type = Cell.TYPE_BOOL
		format_code = NumberFormat.FORMAT_NUMBER		 
	elif field.__class__ is DecimalField:
		data_type = Cell.TYPE_NUMERIC
		format_code = '0.' + '0' * field.decimal_places
	elif field.__class__ is FloatField:
		data_type = Cell.TYPE_NUMERIC
		format_code = NumberFormat.FORMAT_NUMBER_00 
	elif field.__class__ in [CharField, CommaSeparatedIntegerField, EmailField, FilePathField, GenericIPAddressField, IPAddressField, SlugField, URLField]:	
		data_type = Cell.TYPE_STRING
		format_code = NumberFormat.FORMAT_TEXT
	elif field.__class__ is TextField:
		value = value[0:65535]
		data_type = Cell.TYPE_STRING
		format_code = NumberFormat.FORMAT_TEXT
	elif field.__class__ is DateField:
		value = datetime.datetime.combine(value, datetime.time())
		value = SharedDate().datetime_to_julian(date=value)
		data_type = Cell.TYPE_NUMERIC
		format_code = NumberFormat.FORMAT_DATE_YYYYMMDD2
	elif field.__class__ is DateTimeField:
		value = SharedDate().datetime_to_julian(date=value)
		data_type = Cell.TYPE_NUMERIC
		format_code = 'yyy-mm-dd h:mm:ss'
	elif field.__class__ is TimeField:
		value = SharedDate().datetime_to_julian(date=value)
		data_type = Cell.TYPE_NUMERIC
		format_code = NumberFormat.FORMAT_DATE_TIME4
	elif field.__class__ is ForeignKey:
		subobj = value		
		if issubclass(field.rel.to, (Entry, User)):
			value = None
			if subobj is not None:
				value = unicode(subobj)
			data_type = Cell.TYPE_STRING
			format_code = NumberFormat.FORMAT_TEXT
		elif depth > 0:
			value = {}
			for subfield in field.rel.to._meta.fields + field.rel.to._meta.many_to_many:
				if not subfield.auto_created and getattr(subobj, subfield.name) is not None:
					value[subfield.name], junk1, junk2 = format_value_excel(subobj, subfield, depth = depth + 1)			
			data_type = Cell.TYPE_STRING
			format_code = NumberFormat.FORMAT_TEXT
		else:
			value = []
			data_type = []
			format_code = []
			for subfield in field.rel.to._meta.fields + field.rel.to._meta.many_to_many:
				if not subfield.auto_created:
					tmp_val, tmp_data_type, tmp_format_code = format_value_excel(subobj, subfield, depth = depth + 1)
					
					if depth == 0 and isinstance(tmp_val, (list, dict)):
						tmp_val = simplejson.dumps(tmp_val)[1:-1]
					
					value.append(tmp_val)
					data_type.append(tmp_data_type)
					format_code.append(tmp_format_code)
	elif field.__class__ is ManyToManyField:		
		data_type = Cell.TYPE_STRING
		format_code = NumberFormat.FORMAT_TEXT		

		if value is None:
			pass
		elif issubclass(field.rel.to, Entry):
			subobjs = value.all()
			value = []
			for subobj in subobjs:
				value.append(unicode(subobj))
			if depth == 0:
				value = ', '.join(value)
		else:
			subobjs = value.all()
			value = []
			for subobj in subobjs:
				subvalue = {}
				for subfield in field.rel.to._meta.fields + field.rel.to._meta.many_to_many:
					if not subfield.auto_created and getattr(subobj, subfield.name) is not None:
						subvalue[subfield.name], junk1, junk2  = format_value_excel(subobj, subfield, depth = depth + 1)
				value.append(subvalue)
				
			if depth == 0:
				value = simplejson.dumps(value)[1:-1]
	else:
		raise Exception('%s class not supported' % field.__class__.__name__)

	return (value, data_type, format_code, )

def batch_import_from_excel(species_wid, fileName, user):
	models = get_models_in_saving_order()
	
	#read data
	data = read_excel_data(fileName)
	
	#check species WID
	if data.has_key('Species'):
		if len(data['Species']) > 1:
			raise ValidationError('Please define species 1 at a time.')
		species = data['Species'][0]
		if species.has_key('id'):
			species_id = species['id']
		else:
			species_id = None
		species_wid = species['wid']
	else:
		try:
			species = Species.objects.get(wid=species_wid)
			species_id = species.id
		except ObjectDoesNotExist:
			raise ValidationError('Please edit an editing PGDB')
		
	if len(Species.objects.exclude(id=species_id).filter(wid=species_wid)) > 0:
		raise ValidationError('Species WID must be unique.')
		
	#check object WIDs unique
	old_wids_list = []
	if species_id is not None:
		qs = SpeciesComponent.objects.values('wid', 'model_type').filter(species__id=species_id)
		for model in models:
			for obj_data in data[model.__name__]:
				if obj_data['id'] is not None:
					qs = qs.exclude(id=obj_data['id'])
		for x in qs:
			old_wids_list.append(x['wid'])
					
	new_wids_list = []
	for model in models:
		if data.has_key(model.__name__):
			for obj_data in data[model.__name__]:
				new_wids_list.append(obj_data['wid'])
		else:
			data[model.__name__] = []
	
	if len(set(new_wids_list)) < len(new_wids_list):
		raise ValidationError('WIDs must be unique')
	if len(set(new_wids_list) & set(old_wids_list)) > 0:
		raise ValidationError('WIDs not unique: %s' % ', '.join(set(new_wids_list) & set(old_wids_list)))
		
	#get all data for the current species
	all_obj_wids = {species_wid: 'Species'}
	all_obj_data = {species_wid: species}
	all_obj_data_by_model = {'Species': [species]}
	for model_name, model in getModels().iteritems():
		if not issubclass(model, SpeciesComponent):
			continue
		all_obj_data_by_model[model_name] = []
		for obj in data[model_name]:
			all_obj_wids[obj['wid']] = obj['model_type']
			all_obj_data[obj['wid']] = obj
			all_obj_data_by_model[model_name].append(obj)
			
		if species_id is not None:
			qs = model.objects.select_related().filter(species__id=species_id)
			for obj in data[model_name]:
				if obj_data['id'] is not None:
					qs = qs.exclude(id=obj['id'])
			for obj in qs:
				all_obj_wids[obj.wid] = obj.model_type
				all_obj_data[obj.wid] = obj	
				all_obj_data_by_model[model_name].append(obj)
		
	''' validation '''
	errors = []
	
	#fields	
	for model in models:
		for obj_data in data[model.__name__]:
			try:
				obj_data = validate_object_fields(model, obj_data, all_obj_wids, species_wid, obj_data['wid'])
			except ValidationError as error:
				errors.append('%s %s invalid: %s' % (model.__name__, obj_data['wid'], format_error_as_html_list(error)))
	if len(errors) > 0:
		raise ValidationError(format_list_html(errors))	
	
	#objects
	errors = []
	for model in models:
		for obj_data in data[model.__name__]:
			try:
				validate_model_objects(model, 
					obj_data, 
					all_obj_data=all_obj_data, 
					all_obj_data_by_model=all_obj_data_by_model)
			except ValidationError as error:
				errors.append('%s %s invalid: %s' % (model.__name__, obj_data['wid'], format_error_as_html_list(error)))
	if len(errors) > 0:
		raise ValidationError(format_list_html(errors))
	
	#uniqueness
	for model in models:
		try:
			validate_model_unique(model, 
				all_obj_data_by_model[model.__name__], 
				all_obj_data=all_obj_data, 
				all_obj_data_by_model=all_obj_data_by_model)
		except ValidationError as error:
			errors.append('%s invalid: %s' % (model.__name__, format_error_as_html_list(error)))
	if len(errors) > 0:
		raise ValidationError(format_list_html(errors))	
	
	''' save '''
	#commit
	obj_list = {}
	for model in models:
		for obj_data in data[model.__name__]:
			if obj_data['id'] is None:
				obj = model()				
			else:
				obj = model.objects.get(id = obj_data['id'])
			obj_list[obj_data['wid']] = save_object_data(species_wid, obj, obj_data, obj_list, user, save=False, save_m2m=False)
			
	#base
	for model in models:
		for obj_data in data[model.__name__]:
			obj = obj_list[obj_data['wid']]			
			obj_list[obj_data['wid']] = save_object_data(species_wid, obj, obj_data, obj_list, user, save=True, save_m2m=False)
	
	#relations
	for model in models:
		for obj_data in data[model.__name__]:
			obj = obj_list[obj_data['wid']]
			obj_list[obj_data['wid']] = save_object_data(species_wid, obj, obj_data, obj_list, user, save=True, save_m2m=True)
		
def read_excel_data(filename):
	#read workbook
	wb = load_workbook(filename = filename)

	#get models
	models = getModels()

	#initialize errors
	errors = []

	#check headers
	for modelname, model in models.iteritems():
		ws = wb.get_sheet_by_name(name = model._meta.verbose_name_plural[:31])
		if ws is None:
			continue

		iRow1 = 0
		iRow2 = 1
		iCol = -1
		for header, subheaders in get_excel_headers(model):
			if len(subheaders) <= 1:
				iCol += 1
				cell = ws.cell(row=iRow2, column=iCol)
				if cell.value != header:
					errors.append('Expecting worksheet "%s" header %s%s "%s", got "%s"' % (ws.title, get_column_letter(iCol+1), iRow2+1, header, cell.value))
			else:
				cell = ws.cell(row=iRow1, column=iCol+1)
				if cell.value != header:
					errors.append('Expecting worksheet "%s" header %s%s "%s", got "%s"' % (ws.title, get_column_letter(iCol+2), iRow1+1, header, cell.value))
				for subheader in subheaders:
					iCol += 1
					cell = ws.cell(row=iRow2, column=iCol)
					if cell.value != subheader:
						errors.append('Expecting worksheet "%s" header %s%s "%s", got "%s"' % (ws.title, get_column_letter(iCol+1), iRow2+1, subheader, cell.value))

		if (iCol + 1) != ws.get_highest_column():
			errors.append('Expecting worksheet "%s" to have exactly %s columns' % (ws.title, iCol + 1))

	if len(errors) > 0:
		raise ValidationError(format_list_html(errors))

	#get data
	newobjects = {}
	for modelname, model in models.iteritems():
		ws = wb.get_sheet_by_name(name = model._meta.verbose_name_plural[:31])
		if ws is None:
			continue

		fields = getModelDataFields(model)
		
		iCol = -1
		iCol_id = None
		for header, subheaders in get_excel_headers(model):
			if header == 'ID':
				iCol_id = iCol + 1
			iCol += max(1, len(subheaders))

		#data
		newobjects[modelname] = []		
		for iRow in range(2, ws.get_highest_row()):
			id = ws.cell(row=iRow, column=iCol_id).value
			if id == '':
				id is None
			obj = {'id': id, 'model_type': modelname}

			iField = -1
			iCol = -1
			for header, subheaders in get_excel_headers(model):
				iField += 1
				iCol += max(1, len(subheaders))
				field = fields[iField]
				
				if not field.editable or field.auto_created:
					continue
				
				if isinstance(field, (CharField, SlugField, TextField)):
					value = ws.cell(row=iRow, column=iCol).value
					if value is None:
						value = ''
					obj[field.name] = value
				elif isinstance(field, (BooleanField, FloatField, IntegerField, NullBooleanField, PositiveIntegerField)):
					value = ws.cell(row=iRow, column=iCol).value
					if value == '':
						value = None
					obj[field.name] = value
				elif isinstance(field, ForeignKey):
					if issubclass(field.rel.to, Entry):
						value = ws.cell(row=iRow, column=iCol).value
						if value == '':
							value = None
						obj[field.name] = value
					else:
						obj[field.name] = {}
						iCol2 = iCol - max(1, len(subheaders))
						is_blank = True
						for subfield in field.rel.to._meta.fields + field.rel.to._meta.many_to_many:
							if subfield.auto_created:
								continue
							
							iCol2 += 1
							if isinstance(subfield, (CharField, SlugField, TextField)):
								value = ws.cell(row=iRow, column=iCol2).value
								if value is None:
									value = ''
							elif isinstance(subfield, (BooleanField, FloatField, IntegerField, NullBooleanField, PositiveIntegerField)):
								value = ws.cell(row=iRow, column=iCol2).value
								if value == '':
									value = None
							elif isinstance(subfield, ForeignKey) and issubclass(subfield.rel.to, Entry):
								value = ws.cell(row=iRow, column=iCol2).value
								if value == '':
									value = None
							elif isinstance(subfield, ManyToManyField) and issubclass(subfield.rel.to, Entry):
								value = ws.cell(row=iRow, column=iCol2).value
								if value is None or value == '':
									value = []
								else:
									value = value.split(', ')
							elif isinstance(subfield, ForeignKey):
								value = ws.cell(row=iRow, column=iCol2).value
								if value == '':
									value = None
								else:
									value = simplejson.loads(value)
							elif isinstance(subfield, ManyToManyField):
								value = ws.cell(row=iRow, column=iCol2).value
								if value is None or value  == '':
									value = []
								else:
									value = simplejson.loads('[' + value + ']')
							
							obj[field.name][subfield.name] = value
							
							if value is not None and not (isinstance(value, (str, unicode)) and value == '') and not (isinstance(value, list) and len(value) == 0):
								is_blank = False
							
						if is_blank:
							obj[field.name] = None
				elif isinstance(field, ManyToManyField):
					if issubclass(field.rel.to, Entry):
						value = ws.cell(row=iRow, column=iCol).value
						if value is None or value == '':
							obj[field.name] = []
						else:
							obj[field.name] = value.split(', ')
					else:
						value = ws.cell(row=iRow, column=iCol).value
						if value is None or value == '':
							obj[field.name] = []
						else:
							try:
								obj[field.name] = simplejson.loads('[' + value + ']')
							except:
								errors.append('Invalid json at field %s of %s %s at cell %s%s' % (field.name, modelname, obj['wid'], get_column_letter(iCol+1), iRow+1, ))
			
			newobjects[modelname].append(obj)
	
	if len(errors) > 0:
		raise ValidationError(format_list_html(errors))
	
	return newobjects

def format_error_as_html_list(error):
	fullmsg = '<ul>' 
	for key, msg in error.message_dict.iteritems():
		if isinstance(msg, (str, unicode, )):
			msg = [msg]
		
		if key == '__all__':
			fullmsg += '<li>%s</li>' % (' '.join(msg), )
		else:
			fullmsg += '<li>%s: %s</li>' % (key, ' '.join(msg), )
	fullmsg += '</ul>'
	return fullmsg
	
def remove_new_objects(newobjects, newsubobjects):
	for model in get_models_in_saving_order():
		for obj in newobjects[model.__name__]:
			try:
				obj.delete()
			except:
				pass
	for obj in newsubobjects:
		if obj.pk is not None:
			try:
				obj.delete()
			except:
				pass

def objectToQuerySet(obj, model = None):
	if model is None:
		model = obj.__class__
	qs = model.objects.none()
	qs._result_cache.append(obj)
	return qs

def format_field_detail_view(obj, field_name, is_user_anonymous):
	if hasattr(obj, 'get_as_html_%s' % field_name):
		val = getattr(obj, 'get_as_html_%s' % field_name)(is_user_anonymous)
		if isinstance(val, float) and val != 0. and val is not None:		
			return ('%.' + str(int(math.ceil(max(0, -math.log10(abs(val)))+2))) + 'f') % val
		return val
	
	field = obj.__class__._meta.get_field_by_name(field_name)[0]
	value = getattr(obj, field_name)
	
	if isinstance(field, RelatedObject):
		field_model = field.model
		value = value.all()
	elif isinstance(field, ManyToManyField):
		field_model = field.rel.to
		value = value.all()
	elif isinstance(field, ForeignKey):
		field_model = field.rel.to
		
	if isinstance(field, (ManyToManyField, RelatedObject)):		
		if issubclass(field_model, Entry):
			results = []
			for subvalue in value:
				results.append('<a href="%s">%s</a>' % (subvalue.get_absolute_url(), subvalue.wid))
			return format_list_html(results)		
			
		results = []
		for subvalue in value:
			if issubclass(field_model, (EntryBooleanData, )):
				if subvalue.value:
					txt = 'Yes'
				else:
					txt = 'No'
			elif issubclass(field_model, (EntryCharData, EntryFloatData, EntryPositiveFloatData, EntryTextData, )):
				if subvalue.units is None:
					txt = '%s' % (subvalue.value, )
				else:
					txt = '%s (%s)' % (subvalue.value, subvalue.units, )
			else:
				subresults = []
				for subfield in field_model._meta.fields + field_model._meta.many_to_many:
					if subfield.auto_created:
						pass
					elif isinstance(subfield, ManyToManyField) and issubclass(subfield.rel.to, Evidence):
						pass
					else:
						subresults.append('%s: %s' % (subfield.verbose_name, format_field_detail_view(subvalue, subfield.name, is_user_anonymous), ))
				txt = ', '.join(subresults)
						
			if issubclass(field_model, EvidencedEntryData) and len(subvalue.evidence.all()) > 0:			
				txt = txt + (' <a href="javascript:void(0)" onclick="toggleEvidence(this);">(Show evidence)</a><div class="evidence" style="display:none;">%s</div>' % format_evidence(subvalue.evidence.all()))
			results.append('<div>%s</div>' % txt)
		return format_list_html(results)
			
	elif isinstance(field, ForeignKey):
		if issubclass(field_model, Entry):
			if value is not None:
				return '<a href="%s">%s</a>' % (value.get_absolute_url(), value.wid)
			else:
				return ''
		
		if value is None:
			return ''
		
		if issubclass(field_model, (EntryBooleanData, )):
			if value.value:
				txt = 'Yes'
			else:
				txt = 'No'
		elif issubclass(field_model, (EntryCharData, EntryFloatData, EntryPositiveFloatData, EntryTextData, )):
			if value.units is None:
				txt = '%s' % (value.value, )
			else:
				txt = '%s (%s)' % (value.value, value.units, )
		else:
			results = []
			for subfield in field_model._meta.fields + field_model._meta.many_to_many:
				if subfield.auto_created:
					pass
				elif isinstance(subfield, ManyToManyField) and issubclass(subfield.rel.to, Evidence):
					pass
				else:
					results.append('%s: %s' % (subfield.verbose_name, format_field_detail_view(value, subfield.name, is_user_anonymous), ))
			txt = ', '.join(results)
			
		if issubclass(field_model, EvidencedEntryData) and len(value.evidence.all()) > 0:			
			txt = txt + (' <a href="javascript:void(0)" onclick="toggleEvidence(this);">(Show evidence)</a><div class="evidence" style="display:none;">%s</div>' % format_evidence(value.evidence.all()))	
			
		return '<div>%s</div>' % txt
			
	elif (field.choices is not None) and (len(field.choices) > 0) and (not isinstance(field, (BooleanField, NullBooleanField))):
		choices = [x[0] for x in field.choices]
		if value in choices:
			value = field.choices[choices.index(value)][1]
		return value
		
	elif isinstance(value, float) and value != 0 and value is not None:	
		return ('%.' + str(int(math.ceil(max(0, -math.log10(abs(value)))+2))) + 'f') % value
		
	else:
		return value

	return ''
	
def format_field_detail_view_helper():
	pass

def convert_modelobject_to_stdobject(obj, is_user_anonymous=False, ancestors = []):
	model = obj.__class__
	
	if issubclass(model, Entry) and len(ancestors) > 0:
		return obj.wid
	
	objDict = {}
	if issubclass(model, Entry):
		fields = getModelDataFields(model, metadata=not is_user_anonymous)
	else:
		fields = model._meta.fields + model._meta.many_to_many
		
	for field in fields:	
		model_val = getattr(obj, field.name)
		if field.auto_created:
			continue
		
		if isinstance(field, ForeignKey):
			if model_val is None:
				stdobj_val = None
			else:
				if issubclass(field.rel.to, Entry):
					stdobj_val = model_val.wid
				elif issubclass(field.rel.to, User):
					stdobj_val = model_val.username
				else: 
					stdobj_val = convert_modelobject_to_stdobject(model_val, is_user_anonymous, ancestors + [obj])
		elif isinstance(field, ManyToManyField):
			stdobj_val = []
			if issubclass(field.rel.to, Entry):
				for model_subval in model_val.all():
					stdobj_val.append(model_subval.wid)
			elif issubclass(field.rel.to, User):
				for model_subval in model_val.all():
					stdobj_val.append(model_subval.username)
			else:
				for model_subval in model_val.all():
					stdobj_subval = convert_modelobject_to_stdobject(model_subval, is_user_anonymous, ancestors + [obj])
					stdobj_val.append(stdobj_subval)
		elif field.choices is not None and len(field.choices) > 0:
			choices = [x[0] for x in field.choices]
			if model_val in choices:
				stdobj_val = unicode(field.choices[choices.index(model_val)][1])
			else:
				stdobj_val = unicode(model_val)
		else:
			stdobj_val = unicode(model_val)
			
		objDict[field.name] = stdobj_val
	
	for field in model._meta.get_all_related_objects() + model._meta.get_all_related_many_to_many_objects():
		if field.get_accessor_name() == '+':
			continue
			
		objDict[field.get_accessor_name()] = []
		for model_subval in getattr(obj, field.get_accessor_name()).all():
			if model_subval not in ancestors:
				objDict[field.get_accessor_name()].append(convert_modelobject_to_stdobject(model_subval, is_user_anonymous, ancestors + [obj]))
		
	return objDict
	
def convert_modelobject_to_xml(obj, xml_doc, is_user_anonymous, ancestors = []):
	model = obj.__class__
	xml_obj = xml_doc.createElement('object')
	
	if issubclass(model, Entry) and len(ancestors) > 0:
		xml_obj.appendChild(xml_doc.createTextNode(unicode(obj.wid)))
		return xml_obj
	
	xml_field = xml_doc.createElement('field')
	xml_field.setAttribute('name', 'model')
	xml_field.appendChild(xml_doc.createTextNode(unicode(model.__name__)))
	xml_obj.appendChild(xml_field)
	
	if issubclass(model, Entry):
		fields = getModelDataFields(model, metadata=not is_user_anonymous)
	else:
		fields = model._meta.fields + model._meta.many_to_many
		
	for field in fields:
		model_val = getattr(obj, field.name)
		if field.auto_created:
			continue
			
		xml_field = xml_doc.createElement('field')
		xml_field.setAttribute('name', field.name)
		xml_obj.appendChild(xml_field)
		
		if isinstance(field, ForeignKey):
			if model_val is None:
				pass
			else:
				if issubclass(field.rel.to, Entry):
					xml_field.appendChild(xml_doc.createTextNode(unicode(model_val.wid)))
				elif issubclass(field.rel.to, User):
					xml_field.appendChild(xml_doc.createTextNode(unicode(model_val.username)))
				else:
					xml_field.appendChild(convert_modelobject_to_xml(model_val, xml_doc, is_user_anonymous, ancestors + [obj]))
		elif isinstance(field, ManyToManyField):
			if issubclass(field.rel.to, Entry):
				for model_subval in model_val.all():
					doc = xml_doc.createElement('object')
					doc.appendChild(xml_doc.createTextNode(unicode(model_subval.wid)))
					xml_field.appendChild(doc)
			elif issubclass(field.rel.to, User):
				for model_subval in model_val.all():
					doc = xml_doc.createElement('object')
					doc.appendChild(xml_doc.createTextNode(unicode(model_subval.username)))
					xml_field.appendChild(doc)
			else:
				for model_subval in model_val.all():
					xml_field.appendChild(convert_modelobject_to_xml(model_subval, xml_doc, is_user_anonymous, ancestors + [obj]))
		elif field.choices is not None and len(field.choices) > 0:
			choices = [x[0] for x in field.choices]
			if model_val in choices:
				xml_field.appendChild(xml_doc.createTextNode(unicode(field.choices[choices.index(model_val)][1])))
			else:
				xml_field.appendChild(xml_doc.createTextNode(unicode(model_val)))
		else:
			xml_field.appendChild(xml_doc.createTextNode(unicode(model_val)))
			
	for field in model._meta.get_all_related_objects() + model._meta.get_all_related_many_to_many_objects():
		if field.get_accessor_name() == '+':
			continue
			
		xml_field = xml_doc.createElement('field')
		xml_field.setAttribute('name', field.get_accessor_name())
		xml_obj.appendChild(xml_field)
			
		for model_subval in getattr(obj, field.get_accessor_name()).all():
			if model_subval not in ancestors:
				xml_field.appendChild(convert_modelobject_to_xml(model_subval, xml_doc, is_user_anonymous, ancestors + [obj]))
	
	return xml_obj
	
def get_edit_form_fields(species_wid, model, obj=None):
	model_type = model.__name__
	form_fields = []
	initial_values = {}
	
	if issubclass(model, Entry):
		fields = [model._meta.get_field_by_name(x)[0] for x in model._meta.field_list]
	else:
		fields = model._meta.fields + model._meta.many_to_many
	
	for model_field in fields:
		if not model_field.editable or model_field.auto_created:
			continue
		
		type = None
		value = None
		fields = None
		choices = None
		
		if isinstance(model_field, ForeignKey):
			if issubclass(model_field.rel.to, Entry):
				type = 'select'
				if issubclass(model_field.rel.to, SpeciesComponent):
					choices = tuple([(x.wid, x.wid) for x in model_field.rel.to.objects.filter(species__wid=species_wid)])
				else:
					choices = tuple([(x.wid, x.wid) for x in model_field.rel.to.objects.all()])
				if obj is not None and getattr(obj, model_field.name) is not None:
					value = getattr(obj, model_field.name).wid
			else:
				type = 'ForeignKey'
				if obj is not None:
					subobj = getattr(obj, model_field.name)
				else:
					subobj = None
				fields, value = get_edit_form_fields(species_wid, model_field.rel.to, obj=subobj)
				if obj is not None and subobj is None:
					value = None
		elif isinstance(model_field, ManyToManyField):
			if issubclass(model_field.rel.to, Entry):
				type = 'multiselect'				
				if issubclass(model_field.rel.to, SpeciesComponent):
					choices = tuple([(x.wid, x.wid) for x in model_field.rel.to.objects.filter(species__wid=species_wid)])
				else:
					choices = tuple([(x.wid, x.wid) for x in model_field.rel.to.objects.all()])
				if obj is not None:
					value = [x.wid for x in getattr(obj, model_field.name).all()]	
				else:
					value = []
			else:
				type = 'ManyToManyField'
				fields, tmp = get_edit_form_fields(species_wid, model_field.rel.to)
				value = []				
				if obj is not None:			
					for subobj in getattr(obj, model_field.name).all():
						junk, tmp = get_edit_form_fields(species_wid, model_field.rel.to, subobj)
						value.append(tmp)				
		elif isinstance(model_field, TextField):
			type = 'textarea'
			if obj is not None:
				value = getattr(obj, model_field.name)
		elif isinstance(model_field, (BooleanField, NullBooleanField, )):
			type = 'select'
			choices = [
				(True, 'True'),
				(False, 'False'),
				]
			if obj is not None:
				value = getattr(obj, model_field.name)
		elif model_field.choices is not None and len(model_field.choices) >= 1:
			type = 'select'
			choices = model_field.choices
			if obj is not None:
				value = getattr(obj, model_field.name)
		else:
			type = 'text'
			if obj is not None:
				value = getattr(obj, model_field.name)
			
		form_fields.append({
			'type': type, 
			'choices': choices,	
			'fields': fields,
			'verbose_name': model_field.verbose_name, 
			'name': model_field.name})
		
		initial_values[model_field.name] = value
			
	return (form_fields, initial_values)

def get_edit_form_data(model, POST, field_prefix=''):
	#get fields
	if issubclass(model, Entry):
		fields = [model._meta.get_field_by_name(x)[0] for x in model._meta.field_list]
	else:
		fields = model._meta.fields	+ model._meta.many_to_many
		
	#process properties
	data = {}
	is_blank = True
	for field in fields:
		if not field.editable or field.auto_created:
			continue
			
		if isinstance(field, ForeignKey):
			if issubclass(field.rel.to, Entry):
				val = POST.get('%s%s' % (field_prefix, field.name, ), None)
				if val == '':
					val is None
			else:
				val = get_edit_form_data(field.rel.to, POST, field_prefix='%s%s_' % (field_prefix, field.name, ))
		elif isinstance(field, ManyToManyField):
			val = []
			if issubclass(field.rel.to, Entry):
				wids = POST.getlist('%s%s' % (field_prefix, field.name, ), [])				
				for wid in wids:
					if wid is not None and wid != '':
						val.append(wid)
			else:
				keys = POST.dict().keys()
				idxs = []				
				for key in keys:
					match = re.match(r'^%s%s_(\d+)_.+' % (field_prefix, field.name, ), key)
					if match:
						idxs.append(match.group(1))
				idxs = set(idxs)
					
				for idx in idxs:				
					tmp = get_edit_form_data(field.rel.to, POST, field_prefix='%s%s_%s_' % (field_prefix, field.name, idx, ))
					if tmp is not None:
						val.append(tmp)
		else:
			val = POST.get('%s%s' % (field_prefix, field.name, ), None)
			if field.null is True and val == '':
				val = None
			
		data[field.name] = val
		if not (val is None or (isinstance(val, (str, unicode)) and val == '') or (isinstance(val, list) and len(val) == 0)):
			is_blank = False
			
	#return
	if is_blank:
		return None
	return data	
	
def validate_object_fields(model, data, wids, species_wid, entry_wid):
	if issubclass(model, Entry):
		fields = [model._meta.get_field_by_name(x)[0] for x in model._meta.field_list]
	else:
		fields = model._meta.fields	+ model._meta.many_to_many
		
	if issubclass(model, Entry):
		data['species'] = species_wid
		
	errors = {}
	for field in fields:
		if not field.editable or field.auto_created:
			continue
			
		try:			
			if isinstance(field, ForeignKey):
				if issubclass(model, Evidence) and issubclass(field.rel.to, SpeciesComponent):
					data[field.name] = entry_wid
							
				if not data.has_key(field.name):
					data[field.name] = None
					
				if data[field.name] is None:
					if not field.null:
						raise ValidationError('Undefined WID %s' % data[field.name])
				else:
					if issubclass(field.rel.to, Entry):	
						if not wids.has_key(data[field.name]):
							raise ValidationError('Undefined WID %s' % data[field.name])
						if not issubclass(getModel(wids[data[field.name]]), field.rel.to):
							raise ValidationError('Invalid WID %s' % data[field.name])
					else:
						try:
							data[field.name] = validate_object_fields(field.rel.to, data[field.name], wids, species_wid, entry_wid)
						except ValidationError as error:
							tmp = ['%s: %s' % (key, val, ) for key, val in error.message_dict.iteritems()]
							raise ValidationError('<ul><li>%s</li></ul>' % '</li><li>'.join(tmp))
							
			elif isinstance(field, ManyToManyField):
				if not data.has_key(field.name):
					data[field.name] = []
					
				if issubclass(field.rel.to, Entry):
					for wid in data[field.name]:
						if not wids.has_key(wid):
							raise ValidationError('Undefined WID %s' % wid)
						if not issubclass(getModel(wids[wid]), field.rel.to):
							raise ValidationError('Invalid WID %s' % wid)
				else:
					for idx in range(len(data[field.name])):
						try:
							data[field.name][idx] = validate_object_fields(field.rel.to, data[field.name][idx], wids, species_wid, entry_wid)
						except ValidationError as error:
							tmp = ['%s: %s' % (key, val, ) for key, val in error.message_dict.iteritems()]
							raise ValidationError('<ul><li>%s</li></ul>' % '</li><li>'.join(tmp))
			else:				
				if data.has_key(field.name):
					data[field.name] = field.clean(data[field.name], None)
				else:
					if field.default != NOT_PROVIDED:
						data[field.name] = field.clean(field.default, None)
					else:
						data[field.name] = field.clean(None, None)
				
		except ValidationError as error:
			errors[field.name] = ', '.join(error.messages)
	
	if len(errors) > 0:
		raise ValidationError(errors)
		
	return data
	
def validate_model_objects(model, obj_data, all_obj_data=None, all_obj_data_by_model=None):
	parent_list = list(model._meta.get_parent_list())
	parent_list.reverse()
	for parent_model in parent_list + [model]:
		if hasattr(parent_model._meta, 'clean'):
			parent_model._meta.clean(parent_model, obj_data, all_obj_data=all_obj_data, all_obj_data_by_model=all_obj_data_by_model)

def validate_model_unique(model, model_objects_data, all_obj_data=None, all_obj_data_by_model=None):
	if not issubclass(model, SpeciesComponent):
		return
		
	parent_list = list(model._meta.get_parent_list())
	parent_list.reverse()
	
	if all_obj_data is None:
		if model_objects_data[0] is Entry:
			species_wid = model_objects_data[0].species.wid
		else:
			species_wid = model_objects_data[0]['species']
		
		if issubclass(model, SpeciesComponent):
			species_id = Species.objects.values('id').get(wid=species_wid)['id']		
			qs = model.objects.filter(species__id=species_id)
		else:
			qs = model.objects.all()
		for obj in model_objects_data:
			if obj['id'] is not None:
				qs = qs.exclude(id=obj['id'])
		for obj in qs:
			model_objects_data.append(obj)		
	
	for ancestor_model in parent_list + [model]:
		if hasattr(ancestor_model._meta, 'validate_unique'):
			ancestor_model._meta.validate_unique(ancestor_model, model_objects_data, 
				all_obj_data=all_obj_data, all_obj_data_by_model=all_obj_data_by_model)
	
def save_object_data(species_wid, obj, obj_data, obj_list, user, save=False, save_m2m=False):
	model = obj.__class__
	if isinstance(obj, Entry) and obj.pk is None:
		obj.created_user = user
	obj.last_updated_user = user
	
	if issubclass(model, Entry):
		fields = [model._meta.get_field_by_name(x)[0] for x in model._meta.field_list]
	else:
		fields = model._meta.fields	+ model._meta.many_to_many
	
	#regular and foreign key fields
	for field in fields:
		if not field.editable or field.auto_created:
			continue
			
		if isinstance(field, ForeignKey):
			if save:
				if issubclass(field.rel.to, Entry):
					if obj_data[field.name] is not None:
						if obj_list.has_key(obj_data[field.name]):
							tmp = obj_list[obj_data[field.name]]
						elif issubclass(field.rel.to, SpeciesComponent):
							tmp = field.rel.to.objects.get(species__wid=species_wid, wid=obj_data[field.name])
						else:
							tmp = field.rel.to.objects.get(wid=obj_data[field.name])
						setattr(obj, field.name, tmp)
					else:
						setattr(obj, field.name, None)
				else:
					if obj_data[field.name] is not None:
						setattr(obj, field.name, save_object_data(species_wid, field.rel.to(), obj_data[field.name], obj_list, user, save=save, save_m2m=save_m2m))
					else:
						setattr(obj, field.name, None)
		elif isinstance(field, ManyToManyField):
			pass
		else:
			setattr(obj, field.name, obj_data[field.name])
	
	#save
	if save:
		if isinstance(obj, SpeciesComponent):
			obj.species = Species.objects.get(wid=species_wid)
		obj.full_clean()
		obj.save()
	
	#many-to-many fields
	for field in fields:
		if not field.editable or field.auto_created:
			continue
			
		if isinstance(field, ManyToManyField):
			if save_m2m:
				getattr(obj, field.name).clear()
				if issubclass(field.rel.to, Entry):
					for wid in obj_data[field.name]:
						if obj_list.has_key(wid):
							tmp = obj_list[wid]
						elif issubclass(field.rel.to, SpeciesComponent):
							tmp = field.rel.to.objects.get(species__wid=species_wid, wid=wid)
						else:
							tmp = field.rel.to.objects.get(wid=wid)
						getattr(obj, field.name).add(tmp)
				else:
					for sub_obj_data in obj_data[field.name]:
						getattr(obj, field.name).add(save_object_data(species_wid, field.rel.to(), sub_obj_data, obj_list, user, save=save, save_m2m=save_m2m))
	
	#save
	if save:
		obj.full_clean()
		obj.save()
		
	#return obj
	return obj
	
def format_sequence_as_html(sequence, lineLen=50):
	htmlL = ''
	htmlC = ''
	htmlR = ''
	
	for i in range(int(ceil(float(len(sequence)) / float(lineLen)))):
		htmlL += '%s<br/>' % (i * lineLen + 1, )
		htmlC += '%s<br/>' % sequence[i*lineLen:(i + 1) * lineLen]
		htmlR += '%s<br/>' % (min(len(sequence), (i + 1) * lineLen), )
	
	return '<div class="sequence"><div>%s</div><div>%s</div><div>%s</div></div>' % (htmlL, htmlC, htmlR)
	
def readFasta(species_wid, filename, user):
	error_messages = []
	
	f = open(filename, 'r')	
	data = f.readlines()
	f.close()
	
	wid = None
	sequences = {}
	for i in range(len(data)):
		if data[i][0] == '>':
			wid = data[i].strip()[1:]
			sequences[wid] = ''
		else:
			if wid is None:
				return (False, 'File does not match FASTA format.')
			sequences[wid] += data[i].strip()	
	
	#retrieve chromosomes
	for wid, sequence in sequences.iteritems():
		try:
			chr = Chromosome.objects.get(species__wid=species_wid, wid=wid)
		except ObjectDoesNotExist as error:
			error_messages.append(error.message)			
			continue		
	if len(error_messages) > 0:
		return (False, format_list_html(error_messages))
	
	#validate
	for wid, sequence in sequences.iteritems():
		chr = Chromosome.objects.get(species__wid=species_wid, wid=wid)
		chr.sequence = sequence
		try:
			chr.full_clean()
		except ValidationError as error:
			error_messages.append(format_error_as_html_list(error))
			continue
	if len(error_messages) > 0:
		return (False, format_list_html(error_messages))

		
	#save
	for wid, sequence in sequences.iteritems():	
		chr = Chromosome.objects.get(species__wid=species_wid, wid=wid)
		chr.sequence = sequence
		chr.full_clean()
		try:
			chr.save()
		except ValueError as error:
			error_messages.append(error.message)
			continue
	if len(error_messages) > 0:
		return (False, format_list_html(error_messages))
				
	return (True, 'Sequences successfully saved!')
	
def write_bibtex(species, qs):
	refs = {}
	for obj in qs:
		if not isinstance(obj, SpeciesComponent):
			continue
			
		if isinstance(obj, Reference):
			refs[obj.wid] = obj
		for sub_obj in obj.references.all():
			refs[sub_obj.wid] = sub_obj
		for field in obj.__class__._meta.fields + obj.__class__._meta.many_to_many:
			if isinstance(field, ForeignKey) and issubclass(field.rel.to, EvidencedEntryData) and getattr(obj, field.name) is not None:
				for evidence in getattr(obj, field.name).evidence.all():
					for ref in evidence.references.all():
						refs[ref.wid] = ref
			if isinstance(field, ManyToManyField) and issubclass(field.rel.to, EvidencedEntryData):
				for sub_obj in getattr(obj, field.name).all():
					for evidence in sub_obj.evidence.all():
						for ref in evidence.references.all():
							refs[ref.wid] = ref
		
	bibs = []
	wids = refs.keys()
	wids.sort()
	for wid in wids:	
		bibs.append(refs[wid].get_as_bibtex())
		
	now = datetime.datetime.now(tzlocal())
	return \
		  '@preamble {"\n' \
		+ ('\t%s WholeCellKB Bibliography\n' % (species.name, )) \
		+ ('\tGenerated by WholeCellKB on %s at %s\n' % (
			now.isoformat(), 
			settings.ROOT_URL + reverse('public.views.exportData', kwargs={'species_wid': species.wid}),
			)) \
		+ ('\t%s %s %s\n' % (html_to_ascii('&copy;'), now.year, 'Covert Lab, Department of Bioengineering, Stanford University', )) \
		+ '"}\n\n' \
		+ '\n\n'.join(bibs)
	
def get_invalid_objects(species_id, validate_fields=False, validate_objects=True, full_clean=False, validate_unique=False):
	#check that species WIDs unique
	wids = []
	for species in Species.objects.values('wid').all():
		wids.append(species['wid'])
	if len(set(wids)) < len(wids):
		for wid in set(wids):
			del wids[wids.index(wid)]
		return [{
			'type': 'model',
			'name': model_name,
			'url': None,
			'message': 'The following species WIDs are not unique:<ul><li>%s</li></ul>' % '</li></li>'.join(wids),
			}]	
			
	#retrieve objects
	all_obj_wids = {}
	all_obj_data = {}
	all_obj_data_by_model = {}
	
	species = Species.objects.select_related().get(id=species_id)
	species_wid = species.wid
	all_obj_wids[species.wid] = 'Species'
	all_obj_data[species.wid] = species
	all_obj_data_by_model['Species'] = [species]
	
	wids = []
	for model_name, model in getModels(SpeciesComponent).iteritems():
		all_obj_data_by_model[model_name] = []
		for obj in model.objects.select_related().filter(species__id=species_id):
			wids.append(obj.wid)
			all_obj_wids[obj.wid] = model_name
			all_obj_data[obj.wid] = obj
			all_obj_data_by_model[model_name].append(obj)
			
	#check that oobject WIDs unique	
	if len(set(wids)) < len(wids):
		for wid in set(wids):
			del wids[wids.index(wid)]
		return [{
			'type': 'model',
			'name': model_name,
			'url': None,
			'message': 'The following WIDs are not unique:<ul><li>%s</li></ul>' % '</li></li>'.join(wids),
			}]

	#validate
	errors = []
	for model_name, model in getModels().iteritems():
		for obj in all_obj_data_by_model[model_name]:
			junk, obj_data = get_edit_form_fields(species_wid, model, obj=obj)
			obj_data['species'] = species_wid
			try:
				if validate_fields:
					validate_object_fields(model, obj_data, all_obj_wids, species_wid, obj_data['wid'])
				if validate_objects:
					validate_model_objects(model, 
						obj_data, 
						all_obj_data=all_obj_data, 
						all_obj_data_by_model=all_obj_data_by_model)
				
				if full_clean:
					related_errors = {}
					for field in model._meta.fields	+ model._meta.many_to_many:
						if field.editable and not field.auto_created:
							try:
								if isinstance(field, ForeignKey):
									if getattr(obj, field.name) is not None:
										getattr(obj, field.name).full_clean()
								elif isinstance(field, ManyToManyField):
									for sub_obj in getattr(obj, field.name).all():
										sub_obj.full_clean()
							except ValidationError as error:
								related_errors[field.name] = format_error_as_html_list(error)
					if len(related_errors) > 0:
						raise ValidationError(related_errors)

					obj.full_clean()		
			except ValidationError as error:
				errors.append({
					'type': 'object',
					'name': obj.wid,
					'url': obj.get_absolute_url(),
					'message': format_error_as_html_list(error),
					})
		if validate_unique:
			try:
				validate_model_unique(model, 
					all_obj_data_by_model[model_name], 
					all_obj_data=all_obj_data, 
					all_obj_data_by_model=all_obj_data_by_model)
			except ValidationError as error:
				errors.append({
					'type': 'model',
					'name': model_name,
					'url': reverse('public.views.list', kwargs={'species_wid': species.wid, 'model_type': model_name}),
					'message': format_error_as_html_list(error),
					})
		
	return errors
	
class EmpiricalFormula(dict):
	def __init__(self, *args, **kwargs):
		if len(args) > 0 and len(kwargs) > 0:
			raise
		if len(args) > 1:
			raise
			
		if len(args) == 1:
			for match in re.finditer(r'([A-Z][a-z]*)([0-9]*)', args[0]):
				if len(match.group(2)) == 0:
					val = 1
				else:
					val = match.group(2)
				val = float(val)
				if val != 0:
					self[match.group(1)] = val
			
		for key, val in kwargs.iteritems():
			val = float(val)
			if val != 0:
				self[key] = val
		
	def __add__(a, b):
		c = deepcopy(a)
		for key, val in b.iteritems():
			if c.has_key(key):
				c[key] += val
			else:
				c[key] = val
				
			if c[key] == 0:
				del c[key]
		return c
		
	def __sub__(a, b):
		return a + b.__neg__()
		
	def __pos__(a):
		return deepcopy(a)
		
	def __neg__(a):
		b = EmpiricalFormula()
		for key, val in a.iteritems():
			b[key] = -val
		return b
		
	def __mul__(a, b):
		b = float(b)
			
		c = EmpiricalFormula()
		if b != 0:
			for key, val in a.iteritems():
				c[key] = val * b
		return c
		
	def get_molecular_weight(self):
		mw = 0
		for key, val in self.iteritems():
			mw += val * getattr(ELEMENT_MWS, key)
		return mw		
		
	def __str__(self):
		return self.__unicode__()
	
	def __unicode__(self):
		txt = ''
		keys = sorted(self.keys(), cmp=compare_molecular_weight)
		for key in keys:
			val = self[key]
			if val == 0:
				continue
			if float(int(val)) == val:
				val = int(val)
			txt += '%s%s' % (key, val, )
		return txt
		
	def get_as_html(self):
		html = ''
		keys = sorted(self.keys(), cmp=compare_molecular_weight)
		for key in keys:
			val = self[key]
			if val == 0:
				continue
			if float(int(val)) == val:
				val = int(val)
			html += '%s<sub>%s</sub>' % (key, val, )
		return html
		
def compare_molecular_weight(x, y):
	if getattr(ELEMENT_MWS, x) > getattr(ELEMENT_MWS, y):
		return 1
	elif getattr(ELEMENT_MWS, x) < getattr(ELEMENT_MWS, y):
		return -1
	else:
		return 0
		
def draw_molecule(smiles, format='svg', width=200, height=200):
	fid = tempfile.NamedTemporaryFile(suffix = '.svg', delete = False)
	filename = fid.name
	fid.close()
	
	if format == 'svg':
		cmd = '/usr/local/bin/molconvert --smiles "%s" -o "%s" svg:w%s,h%s,transbg' % (smiles, filename, width, height)
		os.system(cmd)
	else:
		os.system('java -classpath .:%s:%s DrawMolecule "%s" %s %s %s "png" "%s"' % (
			'/home/projects/WholeCell/knowledgebase_2/public',
			'/usr/share/ChemAxon/MarvinBeans-5.10.4/lib/MarvinBeans.jar', 
			smiles, 
			width, height, '213 228 246', filename))
	
	fid = open(filename)
	img = fid.read()
	fid.close()
		
	os.remove(filename)
	
	return img
